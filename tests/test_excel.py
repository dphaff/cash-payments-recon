from openpyxl import load_workbook

from cash_recon.excel import write_excel_workbook
from cash_recon.exceptions import classify_all_exceptions
from cash_recon.io.bank_receipts import load_bank_receipts
from cash_recon.io.internal_events import load_internal_events
from cash_recon.io.psp_settlement import load_psp_settlement
from cash_recon.mi import build_mi_summary
from cash_recon.outputs import build_run_output_dir
from cash_recon.recon.internal_psp import reconcile_internal_to_psp
from cash_recon.recon.psp_bank import reconcile_psp_batches_to_bank
from cash_recon.recon.psp_batches import derive_psp_batch_totals


def test_write_excel_workbook_creates_expected_sheets(tmp_path):
    internal_events = load_internal_events("examples/internal_events_valid.csv")
    psp_rows = load_psp_settlement("examples/psp_settlement_valid.csv")
    bank_receipts = load_bank_receipts("examples/bank_receipts_mismatch.csv")
    batch_totals = derive_psp_batch_totals(psp_rows)

    internal_psp_results = reconcile_internal_to_psp(
        internal_events=internal_events,
        psp_rows=psp_rows,
    )

    psp_bank_results = reconcile_psp_batches_to_bank(
        batch_totals=batch_totals,
        bank_receipts=bank_receipts,
    )

    exceptions = classify_all_exceptions(
        internal_psp_results=internal_psp_results,
        psp_bank_results=psp_bank_results,
    )

    mi_summary = build_mi_summary(
        run_id="RUN-TEST",
        internal_events=internal_events,
        psp_rows=psp_rows,
        bank_receipts=bank_receipts,
        batch_totals=batch_totals,
        internal_psp_results=internal_psp_results,
        psp_bank_results=psp_bank_results,
        exceptions=exceptions,
    )

    output_dir = build_run_output_dir(
        outdir=str(tmp_path),
        run_id="RUN-TEST",
    )

    output_path = write_excel_workbook(
        output_dir=output_dir,
        mi_summary=mi_summary,
        internal_psp_results=internal_psp_results,
        psp_bank_results=psp_bank_results,
        exceptions=exceptions,
    )

    assert output_path.exists()

    workbook = load_workbook(output_path)

    assert workbook.sheetnames == [
        "MI Summary",
        "Internal to PSP",
        "PSP to Bank",
        "Exceptions",
    ]


def test_write_excel_workbook_contains_mi_summary_values(tmp_path):
    internal_events = load_internal_events("examples/internal_events_valid.csv")
    psp_rows = load_psp_settlement("examples/psp_settlement_valid.csv")
    bank_receipts = load_bank_receipts("examples/bank_receipts_mismatch.csv")
    batch_totals = derive_psp_batch_totals(psp_rows)

    internal_psp_results = reconcile_internal_to_psp(
        internal_events=internal_events,
        psp_rows=psp_rows,
    )

    psp_bank_results = reconcile_psp_batches_to_bank(
        batch_totals=batch_totals,
        bank_receipts=bank_receipts,
    )

    exceptions = classify_all_exceptions(
        internal_psp_results=internal_psp_results,
        psp_bank_results=psp_bank_results,
    )

    mi_summary = build_mi_summary(
        run_id="RUN-TEST",
        internal_events=internal_events,
        psp_rows=psp_rows,
        bank_receipts=bank_receipts,
        batch_totals=batch_totals,
        internal_psp_results=internal_psp_results,
        psp_bank_results=psp_bank_results,
        exceptions=exceptions,
    )

    output_dir = build_run_output_dir(
        outdir=str(tmp_path),
        run_id="RUN-TEST",
    )

    output_path = write_excel_workbook(
        output_dir=output_dir,
        mi_summary=mi_summary,
        internal_psp_results=internal_psp_results,
        psp_bank_results=psp_bank_results,
        exceptions=exceptions,
    )

    workbook = load_workbook(output_path)
    sheet = workbook["MI Summary"]

    assert sheet["A1"].value == "Metric"
    assert sheet["B1"].value == "Value"
    assert sheet["A2"].value == "Run ID"
    assert sheet["B2"].value == "RUN-TEST"
    assert sheet["A13"].value == "Total exceptions"
    assert sheet["B13"].value == 2