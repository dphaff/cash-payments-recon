from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from cash_recon.exceptions import ReconException
from cash_recon.mi import MISummary
from cash_recon.recon.internal_psp import InternalPSPReconResult
from cash_recon.recon.psp_bank import PSPBankReconResult


HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="D9EAF7",
)

TITLE_FILL = PatternFill(
    fill_type="solid",
    fgColor="EDF4FB",
)


def write_excel_workbook(
    output_dir: Path,
    mi_summary: MISummary,
    internal_psp_results: list[InternalPSPReconResult],
    psp_bank_results: list[PSPBankReconResult],
    exceptions: list[ReconException],
) -> Path:
    output_path = output_dir / "reconciliation_report.xlsx"

    workbook = Workbook()

    mi_sheet = workbook.active
    mi_sheet.title = "MI Summary"

    _write_mi_summary_sheet(mi_sheet, mi_summary)
    _write_internal_psp_sheet(
        workbook=workbook,
        results=internal_psp_results,
    )
    _write_psp_bank_sheet(
        workbook=workbook,
        results=psp_bank_results,
    )
    _write_exceptions_sheet(
        workbook=workbook,
        exceptions=exceptions,
    )

    workbook.save(output_path)

    return output_path


def _write_mi_summary_sheet(sheet, summary: MISummary) -> None:
    rows = [
        ("Metric", "Value"),
        ("Run ID", summary.run_id),
        ("Internal event count", summary.internal_event_count),
        ("PSP row count", summary.psp_row_count),
        ("Bank receipt count", summary.bank_receipt_count),
        ("PSP batch count", summary.psp_batch_count),
        ("Internal to PSP matched", summary.internal_psp_matched_count),
        ("Internal missing in PSP", summary.internal_missing_in_psp_count),
        ("PSP missing in internal", summary.psp_missing_in_internal_count),
        ("PSP to bank matched", summary.psp_bank_matched_count),
        (
            "Expected payout missing in bank",
            summary.expected_payout_missing_in_bank_count,
        ),
        (
            "Bank receipt missing expected payout",
            summary.bank_receipt_missing_expected_payout_count,
        ),
        ("Total exceptions", summary.total_exception_count),
        ("High severity exceptions", summary.high_severity_exception_count),
        ("Medium severity exceptions", summary.medium_severity_exception_count),
        ("Total expected payout amount", str(summary.total_expected_payout_amount)),
        ("Total bank receipt amount", str(summary.total_bank_receipt_amount)),
    ]

    for row in rows:
        sheet.append(row)

    _format_header_row(sheet, row_number=1)
    _autofit_columns(sheet)


def _write_internal_psp_sheet(
    workbook: Workbook,
    results: list[InternalPSPReconResult],
) -> None:
    sheet = workbook.create_sheet("Internal to PSP")

    sheet.append(
        [
            "status",
            "merchant_reference",
            "event_type",
            "gross_amount",
            "internal_event_id",
            "psp_transaction_id",
        ]
    )

    for result in results:
        sheet.append(
            [
                result.status,
                result.merchant_reference,
                result.event_type,
                str(result.gross_amount),
                result.internal_event_id,
                result.psp_transaction_id,
            ]
        )

    _format_header_row(sheet, row_number=1)
    _autofit_columns(sheet)


def _write_psp_bank_sheet(
    workbook: Workbook,
    results: list[PSPBankReconResult],
) -> None:
    sheet = workbook.create_sheet("PSP to Bank")

    sheet.append(
        [
            "status",
            "settlement_batch_id",
            "bank_transaction_id",
            "expected_payout_amount",
            "bank_amount",
            "currency",
        ]
    )

    for result in results:
        sheet.append(
            [
                result.status,
                result.settlement_batch_id,
                result.bank_transaction_id,
                (
                    str(result.expected_payout_amount)
                    if result.expected_payout_amount is not None
                    else None
                ),
                str(result.bank_amount) if result.bank_amount is not None else None,
                result.currency,
            ]
        )

    _format_header_row(sheet, row_number=1)
    _autofit_columns(sheet)


def _write_exceptions_sheet(
    workbook: Workbook,
    exceptions: list[ReconException],
) -> None:
    sheet = workbook.create_sheet("Exceptions")

    sheet.append(
        [
            "exception_type",
            "source_stage",
            "severity",
            "merchant_reference",
            "settlement_batch_id",
            "amount",
            "description",
        ]
    )

    for exception in exceptions:
        sheet.append(
            [
                exception.exception_type,
                exception.source_stage,
                exception.severity,
                exception.merchant_reference,
                exception.settlement_batch_id,
                str(exception.amount) if exception.amount is not None else None,
                exception.description,
            ]
        )

    _format_header_row(sheet, row_number=1)
    _autofit_columns(sheet)


def _format_header_row(sheet, row_number: int) -> None:
    for cell in sheet[row_number]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL

    sheet.freeze_panes = "A2"


def _autofit_columns(sheet) -> None:
    for column_cells in sheet.columns:
        max_length = 0
        column_number = column_cells[0].column
        column_letter = get_column_letter(column_number)

        for cell in column_cells:
            if cell.value is not None:
                value_length = len(str(cell.value))

                if value_length > max_length:
                    max_length = value_length

        adjusted_width = min(max_length + 2, 45)
        sheet.column_dimensions[column_letter].width = adjusted_width